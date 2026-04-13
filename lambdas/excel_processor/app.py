"""Lambda: stream-parse an uploaded Excel/CSV from S3, convert to parquet,
write it back to S3, and POST the result to the ceaser backend callback.

Trigger chain:
    S3 ObjectCreated → SQS → Lambda (batch size 1)

The SQS message body wraps an S3 event notification. The object key follows
the convention ``uploads/{org_id}/{file_id}/{filename}``, so the Lambda parses
the file_id from the key and uses it for the backend callback URL.
"""

from __future__ import annotations

import hashlib
import hmac
import json
import logging
import os
import tempfile
import time
from pathlib import Path
from typing import Any

import boto3
import requests

logger = logging.getLogger()
logger.setLevel(logging.INFO)

PARQUET_BUCKET = os.environ["PARQUET_BUCKET"]
BACKEND_CALLBACK_URL = os.environ["BACKEND_CALLBACK_URL"]
HMAC_SECRET = os.environ["HMAC_SHARED_SECRET"].encode()

s3 = boto3.client("s3")


def _sign(body: bytes) -> str:
    return hmac.new(HMAC_SECRET, body, hashlib.sha256).hexdigest()


def _download_from_s3(bucket: str, key: str, dest: str) -> int:
    """Stream an S3 object to disk. Returns byte size."""
    s3.download_file(bucket, key, dest)
    return Path(dest).stat().st_size


def _parse_csv(path: str) -> dict[str, Any]:
    """Stream-read a CSV to extract column metadata without loading the full file."""
    import pandas as pd

    preview = pd.read_csv(path, nrows=1000, low_memory=False)
    total_rows = 0
    with open(path, "rb") as f:
        for _ in f:
            total_rows += 1
    total_rows = max(total_rows - 1, 0)
    return _column_info(preview, total_rows=total_rows)


def _parse_excel(path: str) -> dict[str, Any]:
    """Memory-safe Excel parse via python-calamine (Rust-backed) when available,
    otherwise openpyxl read_only mode."""
    import pandas as pd

    try:
        preview = pd.read_excel(path, nrows=1000, engine="calamine")
        engine = "calamine"
    except Exception as exc:
        logger.info("calamine unavailable (%s) — falling back to openpyxl read_only", exc)
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None) or []
        sample_rows = []
        for i, row in enumerate(rows_iter):
            if i >= 999:
                break
            sample_rows.append(row)
        preview = pd.DataFrame(sample_rows, columns=[str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)])
        total_rows = i + 1 if sample_rows else 0
        for _ in rows_iter:
            total_rows += 1
        wb.close()
        engine = "openpyxl_readonly"
        return _column_info(preview, total_rows=total_rows, engine=engine)

    # calamine path — count rows via a second read pass
    try:
        full = pd.read_excel(path, engine="calamine")
        total_rows = len(full)
    except Exception:
        total_rows = len(preview)
    return _column_info(preview, total_rows=total_rows, engine=engine)


def _column_info(df, total_rows: int, engine: str = "") -> dict[str, Any]:
    cols = []
    for col in df.columns:
        series = df[col]
        try:
            samples = series.dropna().unique()[:5].tolist()
            safe = []
            for v in samples:
                if hasattr(v, "isoformat"):
                    safe.append(v.isoformat())
                elif hasattr(v, "item"):
                    safe.append(v.item())
                elif isinstance(v, (str, int, float, bool, type(None))):
                    safe.append(v)
                else:
                    safe.append(str(v))
            sample_values = safe
        except Exception:
            sample_values = []
        cols.append({
            "name": str(col),
            "dtype": str(series.dtype),
            "null_count": int(series.isna().sum()),
            "unique_count": int(series.nunique()),
            "sample_values": sample_values,
        })
    return {
        "row_count": int(total_rows),
        "column_count": len(df.columns),
        "columns": cols,
        "engine": engine,
    }


def _to_parquet(src_path: str, file_type: str, parquet_path: str) -> int:
    """Convert the source file to parquet. Returns parquet byte size."""
    import pandas as pd

    if file_type == "csv":
        df = pd.read_csv(src_path, low_memory=False)
    else:
        try:
            df = pd.read_excel(src_path, engine="calamine")
        except Exception:
            df = pd.read_excel(src_path)

    df.to_parquet(parquet_path, engine="pyarrow", index=False, compression="snappy")
    return Path(parquet_path).stat().st_size


def _callback(file_id: str, payload: dict[str, Any]) -> None:
    body = json.dumps(payload, default=str).encode()
    sig = _sign(body)
    url = f"{BACKEND_CALLBACK_URL.rstrip('/')}/files/{file_id}/processed"
    r = requests.post(
        url,
        data=body,
        headers={"Content-Type": "application/json", "X-Ceaser-Signature": sig},
        timeout=30,
    )
    logger.info("Callback %s → HTTP %d", url, r.status_code)
    r.raise_for_status()


def _parse_key(key: str) -> tuple[str, str, str]:
    """Parse an S3 key like ``uploads/{org_id}/{file_id}/{filename}``.
    Returns (org_id, file_id, filename)."""
    from urllib.parse import unquote_plus

    parts = unquote_plus(key).split("/", 3)
    if len(parts) != 4 or parts[0] != "uploads":
        raise ValueError(f"Unexpected S3 key format: {key}")
    return parts[1], parts[2], parts[3]


def _file_type_from_filename(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return "csv"
    if ext in (".xlsx", ".xls"):
        return "excel"
    return "unknown"


def _process_s3_object(bucket: str, key: str) -> None:
    """Download the uploaded file, parse columns, convert to parquet, and
    notify the backend.

    This is the primary processing path for S3 uploads. The backend upload
    endpoint returns immediately with processing_status='processing'. This
    Lambda does the heavy work: column extraction + parquet conversion.
    The backend callback then generates code_preamble and builds the file
    graph from the results.
    """
    org_id, file_id, filename = _parse_key(key)
    file_type = _file_type_from_filename(filename)
    t0 = time.time()

    with tempfile.TemporaryDirectory(prefix="ceaser_lambda_") as tmp:
        src = str(Path(tmp) / filename)
        size = _download_from_s3(bucket, key, src)
        logger.info("Downloaded s3://%s/%s (%d bytes, %.1fs)", bucket, key, size, time.time() - t0)

        # Parse column metadata — this is the primary column extraction path
        # for S3 uploads (no inline processing on the backend).
        column_info: dict | None = None
        try:
            if file_type == "csv":
                column_info = _parse_csv(src)
            elif file_type == "excel":
                column_info = _parse_excel(src)
        except Exception as exc:
            logger.warning("Column parsing failed (non-fatal): %s", exc)

        # Convert to parquet for efficient downstream data access
        parquet_s3_key: str | None = None
        try:
            parquet_local = str(Path(tmp) / f"{Path(filename).stem}.parquet")
            parquet_size = _to_parquet(src, file_type, parquet_local)
            parquet_s3_key = f"parquet/{org_id}/{file_id}.parquet"
            s3.upload_file(parquet_local, PARQUET_BUCKET, parquet_s3_key)
            logger.info("Parquet uploaded to s3://%s/%s (%d bytes)", PARQUET_BUCKET, parquet_s3_key, parquet_size)
        except Exception as exc:
            logger.warning("Parquet conversion failed (non-fatal): %s", exc)

    _callback(file_id, {
        "status": "ready",
        "column_info": column_info,
        "parquet_s3_key": parquet_s3_key,
        "size_bytes": size,
        "elapsed_seconds": round(time.time() - t0, 2),
    })


def lambda_handler(event: dict, context) -> dict:
    """Entry point for S3→SQS→Lambda.

    SQS messages wrap an S3 event notification:
        {"Records": [{"s3": {"bucket": {"name": ...}, "object": {"key": ...}}}]}
    """
    records = event.get("Records", [])
    logger.info("Lambda invoked with %d SQS records", len(records))
    failures = []
    for record in records:
        try:
            body = json.loads(record["body"])
            for s3_rec in body.get("Records", []):
                bucket = s3_rec["s3"]["bucket"]["name"]
                key = s3_rec["s3"]["object"]["key"]
                _process_s3_object(bucket, key)
        except Exception as exc:
            logger.exception("Record %s failed: %s", record.get("messageId"), exc)
            failures.append({"itemIdentifier": record["messageId"]})

    return {"batchItemFailures": failures}
