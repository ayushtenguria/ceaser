"""Formula Extractor Agent — parses formulas from Excel sheets.

Single job: extract formula strings and identify cross-sheet references.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


@dataclass
class FormulaInfo:
    """Extracted formula information."""
    cell: str           # "B5"
    formula: str        # "=SUM(A1:A10)"
    sheet_name: str     # Which sheet this formula is in
    references_sheets: list[str] = field(default_factory=list)  # Cross-sheet refs


@dataclass
class FormulaExtractionResult:
    """Result of extracting all formulas from a workbook."""
    total_formulas: int = 0
    formulas_by_sheet: dict[str, list[FormulaInfo]] = field(default_factory=dict)
    cross_sheet_references: list[tuple[str, str]] = field(default_factory=list)  # (source_sheet, target_sheet)
    has_vlookups: bool = False
    has_index_match: bool = False


def extract_formulas(file_path: str) -> FormulaExtractionResult:
    """Extract all formulas from an Excel file.

    Only works for .xlsx files. CSVs have no formulas.
    Returns empty result for non-xlsx files.
    """
    path = Path(file_path)
    result = FormulaExtractionResult()

    if path.suffix.lower() not in (".xlsx",):
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        sheet_names = set(wb.sheetnames)

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            sheet_formulas: list[FormulaInfo] = []

            max_row = min(ws.max_row or 0, 10000)  # Cap scan
            for row in ws.iter_rows(max_row=max_row):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_str = cell.value

                        # Find cross-sheet references
                        refs = _extract_sheet_references(formula_str, sheet_names)

                        info = FormulaInfo(
                            cell=cell.coordinate,
                            formula=formula_str,
                            sheet_name=ws_name,
                            references_sheets=refs,
                        )
                        sheet_formulas.append(info)

                        # Track cross-sheet refs
                        for ref_sheet in refs:
                            if ref_sheet != ws_name:
                                result.cross_sheet_references.append((ws_name, ref_sheet))

                        # Track function usage
                        upper = formula_str.upper()
                        if "VLOOKUP" in upper:
                            result.has_vlookups = True
                        if "INDEX" in upper and "MATCH" in upper:
                            result.has_index_match = True

            if sheet_formulas:
                result.formulas_by_sheet[ws_name] = sheet_formulas
                result.total_formulas += len(sheet_formulas)

        wb.close()

    except Exception as exc:
        logger.warning("Formula extraction failed: %s", exc)

    # Deduplicate cross-sheet references
    result.cross_sheet_references = list(set(result.cross_sheet_references))

    logger.info("Extracted %d formulas, %d cross-sheet refs",
                result.total_formulas, len(result.cross_sheet_references))
    return result


def _extract_sheet_references(formula: str, known_sheets: set[str]) -> list[str]:
    """Extract sheet names referenced in a formula."""
    refs: list[str] = []

    # Pattern: Sheet1!B5, 'Sheet Name'!B5
    matches = re.findall(r"['\"]?(\w[\w\s]*?)['\"]?\![A-Z]+\d+", formula)
    for match in matches:
        sheet = match.strip("'\"")
        if sheet in known_sheets:
            refs.append(sheet)

    return list(set(refs))
