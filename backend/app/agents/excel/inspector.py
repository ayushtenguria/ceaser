"""Workbook Inspector Agent — opens a file and reports what's inside.

Single job: inspect file metadata WITHOUT loading full data.
Fast — reads only metadata, not all rows.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)


@dataclass
class SheetInfo:
    """Metadata about one sheet (without loading data)."""
    name: str
    estimated_rows: int = 0
    estimated_cols: int = 0
    has_formulas: bool = False
    has_merged_cells: bool = False


@dataclass
class WorkbookInspection:
    """Result of inspecting a workbook."""
    file_name: str
    file_path: str
    file_size_bytes: int
    file_type: str
    sheet_count: int = 0
    sheets: list[SheetInfo] = field(default_factory=list)
    encoding: str = "utf-8"
    delimiter: str = ","
    warnings: list[str] = field(default_factory=list)


def inspect_workbook(file_path: str) -> WorkbookInspection:
    """Inspect a file and return metadata WITHOUT loading all data.

    Fast operation — reads only headers and metadata.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    result = WorkbookInspection(
        file_name=path.name,
        file_path=str(path),
        file_size_bytes=path.stat().st_size,
        file_type=path.suffix.lower().lstrip("."),
    )

    if result.file_type == "csv":
        result = _inspect_csv(path, result)
    elif result.file_type in ("xlsx", "xls"):
        result = _inspect_excel(path, result)
    else:
        result.warnings.append(f"Unknown file type: {result.file_type}")

    logger.info("Inspected %s: %d sheets, type=%s", path.name, result.sheet_count, result.file_type)
    return result


def _inspect_csv(path: Path, result: WorkbookInspection) -> WorkbookInspection:
    """Inspect a CSV file."""
    result.encoding = _detect_encoding(path)

    result.delimiter = _detect_delimiter(path, result.encoding)

    try:
        row_count = 0
        with open(path, "r", encoding=result.encoding, errors="replace") as f:
            for _ in f:
                row_count += 1
        row_count = max(0, row_count - 1)
    except Exception:
        row_count = 0

    try:
        df_head = pd.read_csv(path, encoding=result.encoding, sep=result.delimiter,
                               nrows=0, on_bad_lines="skip")
        col_count = len(df_head.columns)
    except Exception:
        col_count = 0

    result.sheet_count = 1
    result.sheets = [SheetInfo(
        name=path.stem,
        estimated_rows=row_count,
        estimated_cols=col_count,
    )]
    return result


def _inspect_excel(path: Path, result: WorkbookInspection) -> WorkbookInspection:
    """Inspect an Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append(SheetInfo(
                name=name,
                estimated_rows=ws.max_row or 0,
                estimated_cols=ws.max_column or 0,
            ))

        wb.close()
        result.sheet_count = len(sheets)
        result.sheets = sheets

        try:
            wb2 = openpyxl.load_workbook(path, read_only=True, data_only=False)
            for name in wb2.sheetnames:
                ws2 = wb2[name]
                for row in ws2.iter_rows(max_row=min(ws2.max_row or 0, 20), values_only=True):
                    for cell in row:
                        if isinstance(cell, str) and cell.startswith("="):
                            for s in sheets:
                                if s.name == name:
                                    s.has_formulas = True
                            break
                    if any(s.has_formulas for s in sheets if s.name == name):
                        break
            wb2.close()
        except Exception:
            pass

    except Exception as exc:
        result.warnings.append(f"Failed to inspect: {exc}")
        logger.warning("Excel inspection failed: %s", exc)

    return result


def _detect_encoding(path: Path) -> str:
    """Detect file encoding."""
    try:
        import chardet
        with open(path, "rb") as f:
            raw = f.read(10240)
        detected = chardet.detect(raw)
        enc = detected.get("encoding", "utf-8") or "utf-8"
        with open(path, "r", encoding=enc) as f:
            f.read(1024)
        return enc
    except Exception:
        for enc in ["utf-8", "latin-1", "cp1252"]:
            try:
                with open(path, "r", encoding=enc) as f:
                    f.read(1024)
                return enc
            except Exception:
                continue
        return "utf-8"


def _detect_delimiter(path: Path, encoding: str) -> str:
    """Detect CSV delimiter."""
    import csv
    try:
        with open(path, "r", encoding=encoding, errors="replace") as f:
            sample = f.read(8192)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
        return dialect.delimiter
    except Exception:
        return ","
