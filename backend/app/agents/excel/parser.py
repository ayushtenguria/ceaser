"""Excel Parser Agent — extracts clean DataFrames from messy Excel files.

Handles: multi-sheet workbooks, offset headers, merged cells, formulas,
large files (500K+ rows), CSV fallback, type detection.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)

_MAX_ROWS = 2_000_000  # Hard limit — beyond this, suggest DB connection
_SAMPLE_SIZE = 3_000   # Rows to sample for type detection / profiling
_HEADER_SCAN_ROWS = 20 # Rows to scan for header detection


@dataclass
class SheetResult:
    """Parsed result for a single sheet."""
    name: str
    df: pd.DataFrame
    row_count: int = 0
    column_count: int = 0
    column_types: dict[str, str] = field(default_factory=dict)
    formulas: dict[str, str] = field(default_factory=dict)
    original_header_row: int = 0
    has_merged_cells: bool = False
    sample_values: dict[str, list] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def __post_init__(self):
        self.row_count = len(self.df)
        self.column_count = len(self.df.columns)


@dataclass
class WorkbookResult:
    """Parsed result for an entire workbook."""
    file_name: str
    file_path: str
    file_size_bytes: int = 0
    sheets: list[SheetResult] = field(default_factory=list)
    total_rows: int = 0
    total_columns: int = 0
    parse_warnings: list[str] = field(default_factory=list)

    def __post_init__(self):
        self.total_rows = sum(s.row_count for s in self.sheets)
        self.total_columns = max((s.column_count for s in self.sheets), default=0)


def parse_workbook(file_path: str) -> WorkbookResult:
    """Parse an Excel or CSV file into clean DataFrames.

    Main entry point. Handles:
    - .xlsx/.xls with multiple sheets
    - .csv with auto-detected delimiter/encoding
    - Large files via chunked reading
    - Messy headers, merged cells, type detection
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    file_size = path.stat().st_size
    suffix = path.suffix.lower()

    result = WorkbookResult(
        file_name=path.name,
        file_path=str(path),
        file_size_bytes=file_size,
    )

    if suffix == ".csv":
        sheet = _parse_csv(path)
        result.sheets = [sheet]
    elif suffix in (".xlsx", ".xls"):
        result.sheets = _parse_excel(path, file_size)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    # Post-process all sheets
    for sheet in result.sheets:
        sheet.column_types = _detect_column_types(sheet.df)
        sheet.sample_values = _extract_sample_values(sheet.df)
        sheet.row_count = len(sheet.df)
        sheet.column_count = len(sheet.df.columns)

    result.total_rows = sum(s.row_count for s in result.sheets)
    result.total_columns = max((s.column_count for s in result.sheets), default=0)

    logger.info(
        "Parsed %s: %d sheets, %d total rows",
        path.name, len(result.sheets), result.total_rows,
    )
    return result


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _parse_excel(path: Path, file_size: int) -> list[SheetResult]:
    """Parse all sheets from an Excel workbook."""
    sheets: list[SheetResult] = []

    # Get sheet names
    xl = pd.ExcelFile(path)
    sheet_names = xl.sheet_names

    for sheet_name in sheet_names:
        try:
            sheet = _parse_single_sheet(path, sheet_name, file_size)
            if sheet is not None:
                sheets.append(sheet)
        except Exception as exc:
            logger.warning("Failed to parse sheet '%s': %s", sheet_name, exc)

    xl.close()
    return sheets


def _parse_single_sheet(path: Path, sheet_name: str, file_size: int) -> SheetResult | None:
    """Parse a single sheet from an Excel file."""
    # First pass: read raw to detect header row
    try:
        raw_df = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=_HEADER_SCAN_ROWS)
    except Exception as exc:
        logger.warning("Cannot read sheet '%s': %s", sheet_name, exc)
        return None

    if raw_df.empty or raw_df.shape[1] < 1:
        logger.info("Skipping empty sheet: %s", sheet_name)
        return None

    # Detect header row
    header_row = _detect_header_row(raw_df)

    # Read the full sheet with correct header
    use_chunked = file_size > 50 * 1024 * 1024  # >50MB

    if use_chunked:
        df = _read_large_sheet(path, sheet_name, header_row)
    else:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)

    if df.empty:
        return None

    # Clean up
    df = _clean_dataframe(df)

    # Check row limit
    warnings = []
    if len(df) > _MAX_ROWS:
        warnings.append(f"Sheet truncated from {len(df)} to {_MAX_ROWS} rows")
        df = df.head(_MAX_ROWS)

    # Extract formulas (only for xlsx)
    formulas = {}
    if path.suffix.lower() == ".xlsx":
        formulas = _extract_formulas(path, sheet_name)

    return SheetResult(
        name=sheet_name,
        df=df,
        original_header_row=header_row,
        has_merged_cells=False,  # Updated by formula extraction
        formulas=formulas,
        warnings=warnings,
    )


def _read_large_sheet(path: Path, sheet_name: str, header_row: int) -> pd.DataFrame:
    """Read a large Excel sheet using openpyxl read-only mode for memory efficiency."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        rows = []
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < header_row:
                continue
            if i == header_row:
                header = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                continue
            if i > _MAX_ROWS + header_row:
                break
            rows.append(row)

        wb.close()

        if not header or not rows:
            return pd.DataFrame()

        return pd.DataFrame(rows, columns=header)
    except Exception as exc:
        logger.warning("Chunked read failed, falling back: %s", exc)
        return pd.read_excel(path, sheet_name=sheet_name, header=header_row, nrows=_MAX_ROWS)


# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------

def _parse_csv(path: Path) -> SheetResult:
    """Parse a CSV file with auto-detected delimiter and encoding."""
    import csv

    # Detect encoding
    encoding = _detect_encoding(path)

    # Detect delimiter
    delimiter = _detect_delimiter(path, encoding)

    # Read CSV
    try:
        df = pd.read_csv(
            path,
            encoding=encoding,
            sep=delimiter,
            on_bad_lines="skip",
            low_memory=False,
            nrows=_MAX_ROWS,
        )
    except Exception:
        # Fallback: try with defaults
        df = pd.read_csv(path, on_bad_lines="skip", nrows=_MAX_ROWS)

    df = _clean_dataframe(df)

    return SheetResult(
        name=path.stem,  # filename without extension
        df=df,
    )


def _detect_encoding(path: Path) -> str:
    """Detect file encoding by reading first 10KB."""
    try:
        import chardet
        with open(path, "rb") as f:
            raw = f.read(10240)
        result = chardet.detect(raw)
        return result.get("encoding", "utf-8") or "utf-8"
    except ImportError:
        return "utf-8"


def _detect_delimiter(path: Path, encoding: str) -> str:
    """Detect CSV delimiter using csv.Sniffer."""
    import csv
    try:
        with open(path, "r", encoding=encoding, errors="replace") as f:
            sample = f.read(8192)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|;")
        return dialect.delimiter
    except Exception:
        return ","


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _detect_header_row(raw_df: pd.DataFrame) -> int:
    """Detect which row is the header by scoring each row.

    Scoring criteria:
    - % of cells that are non-null strings (higher = more likely header)
    - % of cells that are unique (headers usually unique)
    - No numeric-only cells (headers are usually text)
    """
    best_row = 0
    best_score = -1

    for i in range(min(len(raw_df), _HEADER_SCAN_ROWS)):
        row = raw_df.iloc[i]
        non_null = row.notna().sum()
        if non_null == 0:
            continue

        # Score components
        str_count = sum(1 for v in row if isinstance(v, str) and len(str(v).strip()) > 0)
        unique_count = row.nunique()
        total = len(row)

        str_ratio = str_count / total if total > 0 else 0
        unique_ratio = unique_count / total if total > 0 else 0
        non_null_ratio = non_null / total if total > 0 else 0

        # Penalize rows that are mostly numeric (data rows, not headers)
        numeric_count = sum(1 for v in row if isinstance(v, (int, float)) and not isinstance(v, bool))
        numeric_penalty = numeric_count / total if total > 0 else 0

        score = (str_ratio * 3) + (unique_ratio * 2) + (non_null_ratio * 1) - (numeric_penalty * 2)

        if score > best_score:
            best_score = score
            best_row = i

    return best_row


# ---------------------------------------------------------------------------
# Data cleaning
# ---------------------------------------------------------------------------

def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Clean a DataFrame: fix column names, drop empty rows/cols."""
    if df.empty:
        return df

    # Clean column names
    df.columns = _clean_column_names(df.columns)

    # Drop completely empty rows and columns
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")

    # Strip whitespace from string columns
    for col in df.select_dtypes(include=["object"]).columns:
        try:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"nan": None, "None": None, "": None})
        except Exception:
            pass

    # Reset index
    df = df.reset_index(drop=True)

    return df


def _clean_column_names(columns: pd.Index) -> list[str]:
    """Clean column names: lowercase, strip, replace special chars."""
    clean = []
    seen: dict[str, int] = {}

    for col in columns:
        name = str(col).strip().lower()
        # Replace special characters with underscore
        name = re.sub(r"[^\w\s]", "_", name)
        name = re.sub(r"\s+", "_", name)
        name = re.sub(r"_+", "_", name)
        name = name.strip("_")

        if not name or name == "nan":
            name = "unnamed"

        # Deduplicate
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0

        clean.append(name)

    return clean


# ---------------------------------------------------------------------------
# Type detection
# ---------------------------------------------------------------------------

def _detect_column_types(df: pd.DataFrame) -> dict[str, str]:
    """Detect column types: numeric, date, currency, percentage, string."""
    types: dict[str, str] = {}

    # Sample for large DataFrames
    sample = df.head(_SAMPLE_SIZE) if len(df) > _SAMPLE_SIZE else df

    for col in sample.columns:
        series = sample[col].dropna()
        if series.empty:
            types[col] = "empty"
            continue

        # Already numeric
        if pd.api.types.is_numeric_dtype(series):
            types[col] = "numeric"
            continue

        # Already datetime
        if pd.api.types.is_datetime64_any_dtype(series):
            types[col] = "date"
            continue

        str_vals = series.astype(str)

        # Check for currency ($1,234.56)
        currency_pattern = "^[\\$€£¥]?\\s*[\\d,]+\\.?\\d*$"
        try:
            currency_match = str_vals.str.match(currency_pattern).sum() / len(str_vals)
        except Exception:
            currency_match = 0
        if currency_match > 0.7:
            types[col] = "currency"
            continue

        # Check for percentage (45%, 0.45)
        pct_pattern = "^\\d+\\.?\\d*\\s*%$"
        try:
            pct_match = str_vals.str.match(pct_pattern).sum() / len(str_vals)
        except Exception:
            pct_match = 0
        if pct_match > 0.7:
            types[col] = "percentage"
            continue

        # Try numeric conversion
        numeric_converted = pd.to_numeric(str_vals.str.replace(",", ""), errors="coerce")
        if numeric_converted.notna().sum() / len(str_vals) > 0.8:
            types[col] = "numeric"
            continue

        # Try date conversion
        try:
            date_converted = pd.to_datetime(str_vals, errors="coerce", infer_datetime_format=True)
            if date_converted.notna().sum() / len(str_vals) > 0.6:
                types[col] = "date"
                continue
        except Exception:
            pass

        types[col] = "string"

    return types


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_sample_values(df: pd.DataFrame, n: int = 5) -> dict[str, list]:
    """Extract sample unique values per column for LLM context."""
    samples: dict[str, list] = {}
    for col in df.columns:
        try:
            unique = df[col].dropna().unique()[:n]
            samples[col] = [
                v.item() if hasattr(v, "item") else str(v) for v in unique
            ]
        except Exception:
            samples[col] = []
    return samples


def _extract_formulas(path: Path, sheet_name: str) -> dict[str, str]:
    """Extract formulas from an xlsx sheet (not computed values)."""
    formulas: dict[str, str] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb[sheet_name]

        for row in ws.iter_rows(max_row=min(ws.max_row or 0, 100)):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[cell.coordinate] = cell.value

        wb.close()
    except Exception as exc:
        logger.debug("Formula extraction failed for %s: %s", sheet_name, exc)

    return formulas
